# Header row
# Code by Avery Cheng
# import necessary features
from re import A
#import xlrd
import random
from openpyxl import Workbook,load_workbook
import time
from gtts import gTTS
import playsound

#initialise variables
#headings={}; tagCol=0; tags={}

# select subject
#xlFile = xlrd.open_workbook("VocabBook.xls")
#options = xlFile.sheet_names()
#vcList = xlFile.sheet_by_index(0)

wb = load_workbook('VocabBook.xlsx')

ws = wb.active

wb.save('VocabBookCodeBackup.xlsx')

words = []

#words should be [Row, Kanji, Hiragana, English, amounts]

initial_word_amounts = {}

row_num = 0
for row in ws.values:
    row_num +=1
    if (row[1] == None):
        print ("sussy")
        break
    else:
        if row[3] == None:
            ws[str('D'+str(row_num))] = 0
            wb.save('VocabBook.xlsx')
        wordrow = [row_num]
        for lad in row[0:3]:
            wordrow.append(lad)
        if row[3] == None:
            wordrow.append(0)
        else:
            wordrow.append(row[3])
        words.append(wordrow)
        initial_word_amounts[row_num] = wordrow[4]


#removing the first lad from the lads
words.pop(0)
initial_word_amounts.pop(1)

#print(words)

print("Welcome to Vocab Book Master!!! Original code by Avery Cheng :)")
print('')
print('Vocab Book:')
for i in words:
    print (i)

print()

# if you wna do the stagger mode
staggeryn = True

repeat = 0
confirmed = False
while not confirmed:
    print('select your repeating mode')
    print()
    print("Kanji to Hiragana to English:")
    print('1 - do not repeat words')
    print('2 - repeat all words')
    print('3 - repeat only words I get wrong')
    print()
    print("English to Hiragana to Kanji:")
    print('4 - do not repeat words')
    print('5 - repeat all words')
    print('6 - repeat only words I get wrong')
    print()
    inp = int(input("repeating mode? "))
    if inp == 1 or inp == 2 or inp == 3 or inp == 4 or inp == 5 or inp == 6:
        repeat = inp
        confirmed = True
    else:
        print("invalid input (input '1','2','3','4','5','6')")
#        confirmed = False

#print(words)

print("")
print("=======================End of setup=======================")
print("")

# -----------------------End of setup-----------------------

#Ideas:
#No Repeats (pop?)

# -----------------------Begin main-------------------------

#def selectQuestion(wordList):
#    word = wordList.pop(random.randrange(len(wordList)))
#    del word[tagCol]
#    answer = [] ; question = []
#    for i in omittedHeadings:
#        answer.append(str(word[(omittedHeadings[i])].value))
#    for i in nomittedHeadings:
#        question.append(str(word[(nomittedHeadings[i])].value))

#    print(omittedHeadings)
#    print(nomittedHeadings)
#    return answer, question

#def selectQuestionRepeat(wordList):
#    word = wordList[random.randrange(len(wordList))].copy()
#    del word[tagCol]
#    answer = [] ; question = []
#    for i in omittedHeadings:
#        answer.append(word.pop(omittedHeadings[i]).value)
#    for i in word:
#        question.append(i.value)
#    return answer, question

def stagger_question(repeat, wordlist, current_amounts):
    word, new_amounts = weighted_rand(wordlist, current_amounts)
    if repeat == 1 or repeat == 4:
        wordlist.remove(word)
    disRow = 0; dis1 = ''; dis2 = ''; dis3 = ''
    disRow = word[0]
    if repeat == 1 or repeat == 2 or repeat == 3:
        dis1 = str(word[1])
        dis2 = str(word[2])
        dis3 = str(word[3])
    elif repeat == 4 or repeat == 5 or repeat == 6:
        dis1 = str(word[3])
        dis2 = str(word[2])
        dis3 = str(word[1])
    if repeat == 1 or repeat == 4:
        new_amounts.pop(disRow)
    current_amounts = new_amounts
    return disRow, dis1, dis2, dis3, current_amounts

def weighted_rand(wordlist, current_amounts):
    weightings = []
    for amount in current_amounts:
        weightings.append((2/3)**current_amounts[amount])
    word_as_a_list = random.choices(wordlist,weights=weightings, k=1)
    word = word_as_a_list[0]
    return word, current_amounts

def say_word(word,langu):
    tts = gTTS(text=word, lang=langu)
    tts.save('__pycache/temp.mp3')
    playsound.playsound("C:\\Users\\chenga\\OneDrive - Kardinia International College\\Documents\\0Work\\2022\\Japanese\\__pycache\\temp.mp3")

finished = False
#dis1, dis2 and dis3 for the staggered
current_amounts = {}
for existing in initial_word_amounts:
        current_amounts[existing] = initial_word_amounts[existing]
        
while True:

    #if staggeryn == True:

    disRow, dis1, dis2, dis3, current_amounts = stagger_question(repeat, words, current_amounts)
    if repeat == 1 or repeat == 2 or repeat == 3:
        if "(" in dis1:
            dis1 = dis2
    elif repeat == 4 or repeat == 5 or repeat == 6:
        if "(" in dis3:
            dis3 = dis2
    if repeat == 1 or repeat == 2 or repeat == 3:
        print(dis1)
        print()
        say_word(dis1, 'ja')
        time.sleep(0.7)
        say_word(dis1, 'ja')
        time.sleep(0.7)
        print(dis3)
        print()
        print("======================")
        print()
        say_word(dis3, 'en')
        time.sleep(1.5)
    elif repeat == 4 or repeat == 5 or repeat == 6:
        print(dis1)
        print()
        say_word(dis1, 'en')
        time.sleep(1)
        print(dis3)
        print()
        print("======================")
        print()
        say_word(dis3, 'ja')
        time.sleep(1.5)

    #adding 1 to the amounts if y is input hee hee
    inp == 'hee hoo'
    
    if inp == 'y':
        prev_amounts = ws[str('D'+str(disRow))].value
        ws[str('D'+str(disRow))] = ws[str('D'+str(disRow))].value + 1
        wb.save('VocabBook.xlsx')
        if repeat == 3:
            words.remove([disRow, dis1, dis2, dis3, current_amounts[disRow]])
            current_amounts.pop(disRow)
        if repeat == 6:
            words.remove([disRow, dis3, dis2, dis1, current_amounts[disRow]])
            current_amounts.pop(disRow)
            
    
    #re getting the swaggy amounts from the big lad list lad thingo lad yes
    row_num = 0
    for row in ws.values:
        row_num +=1
        if row[1] != None:
            initial_word_amounts[row_num] = row[3]
    for existing in current_amounts:
        current_amounts[existing] = initial_word_amounts[existing]
    
    current_amounts[1] = 'peepee poopoo'
    current_amounts.pop(1)

    if finished == True:
        break

    if len(current_amounts) < 2:
        finished = True

print()
print ("good study! have a good day :)")
input()
    
