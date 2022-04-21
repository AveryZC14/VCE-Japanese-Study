# Header row
# Original code by fig
# modifications made by andre
#swag from both of us
# import necessary features
from re import A
#import xlrd
import random
from openpyxl import Workbook,load_workbook
from math import ceil

#initialise variables
#headings={}; tagCol=0; tags={}

# select subject
#xlFile = xlrd.open_workbook("VocabBook.xls")
#options = xlFile.sheet_names()
#vcList = xlFile.sheet_by_index(0)

wb = load_workbook('VocabBook.xlsx')

ws = wb.active

wb.save('VocabBookCodeBackup.xlsx')

words = []

initial_word_amounts_je = {}
initial_word_amounts_ej = {}

row_num = 0
for row in ws.values:
    row_num +=1
    if (row[1] == None):
        print ("sussy")
        break
    else:
        if row[0] == None:
            ws[str('A'+str(row_num))] = row[1]
            wb.save('VocabBook.xlsx')
        if row[3] == None:
            ws[str('D'+str(row_num))] = 0
            wb.save('VocabBook.xlsx')
        if row[4] == None:
            ws[str('E'+str(row_num))] = 0
            wb.save('VocabBook.xlsx')
        wordrow = [row_num]
        if row[0] == None:
            wordrow.append(row[1])
        else:
            wordrow.append(row[0])
            
        for lad in row[1:3]:
            wordrow.append(lad)
        if row[3] == None:
            wordrow.append(0)
        else:
            wordrow.append(row[3])
        if row[4] == None:
            wordrow.append(0)
        else:
            wordrow.append(row[4])
        words.append(wordrow)
        initial_word_amounts_je[row_num] = wordrow[4]
        initial_word_amounts_ej[row_num] = wordrow[5]


#removing the first lad from the lads
words.pop(0)
initial_word_amounts_je.pop(1)
initial_word_amounts_ej.pop(1)

#print(words)

print("Welcome to Vocab Book Master!!! Original code by Fionn Mitchell and Andrew Cheng :)")
print('')
print('Vocab Book:')
for i in words:
    print (i)

print("")
print("==========================End of data dump==========================")
print("")

# if you wna do the stagger mode
staggeryn = True

repeat = 0
confirmed = False
while not confirmed:
    print("select your repeating mode or type 'help' for some help")
    print()
    print("Kanji to Hiragana to English:")
    print('1 - do not repeat words')
    print('2 - repeat all words')
    print('3 - repeat only words I get wrong')
    print()
    print("English to Hiragana to Kanji:")
    print('4 - do not repeat words')
    print('5 - repeat all words')
    print('6 - repeat only words I get wrong')
    print()
    inp = input("repeating mode? ")
    try:
        inp = int(inp)    
        if inp == 1 or inp == 2 or inp == 3 or inp == 4 or inp == 5 or inp == 6:
            repeat = inp
            confirmed = True
        else:
            print()
            print(':(')
            print()
            print("invalid input (number too high what u doin)")
    except:
        if inp == 'help':
            print("")
            print("==========================End of data dump==========================")
            print("")
            print('the code will spit vocab at ya')
            print("the higher the word's score, the less likely it is to pop up")
            print('this ensures you practice the words you know the least well every time')
            print("press enter to go to the next part of the word (kanji, hiragana, english) or (english, hiragana, kanji)")
            print("at the end of each word, the code will prompt you for how you went")
            print()
            print("type 'y' or 'z' if you got it right, this will increase that word's score by 1")
            print()
            print("type 'x' or 'n' if you got it wrong and would like to work on it, this will decrease that word's score by 1")
            print()
            print("leave it blank or type anything else if you got it wrong but it's aight, this will not affect that word's score")
            print()
            print(':)')
            print()
        else:
            print()
            print(':(')
            print()
            print("invalid input (input '1','2','3','4','5','6' or 'help' )")
    #       confirmed = False

#print(words)

print("")
print("==========================End of setup==========================")
print("")

# -----------------------End of setup-----------------------

#Ideas:
#No Repeats (pop?)

# -----------------------Begin main-------------------------

#def selectQuestion(wordList):
#    word = wordList.pop(random.randrange(len(wordList)))
#    del word[tagCol]
#    answer = [] ; question = []
#    for i in omittedHeadings:
#        answer.append(str(word[(omittedHeadings[i])].value))
#    for i in nomittedHeadings:
#        question.append(str(word[(nomittedHeadings[i])].value))

#    print(omittedHeadings)
#    print(nomittedHeadings)
#    return answer, question

#def selectQuestionRepeat(wordList):
#    word = wordList[random.randrange(len(wordList))].copy()
#    del word[tagCol]
#    answer = [] ; question = []
#    for i in omittedHeadings:
#        answer.append(word.pop(omittedHeadings[i]).value)
#    for i in word:
#        question.append(i.value)
#    return answer, question

def stagger_question(repeat, wordlist, current_amounts):
    word, new_amounts = weighted_rand(wordlist, current_amounts)
    if repeat == 1 or repeat == 4:
        wordlist.remove(word)
    dis_row = 0; dis1 = ''; dis2 = ''; dis3 = ''
    dis_row = word[0]
    if repeat == 1 or repeat == 2 or repeat == 3:
        dis1 = str(word[1])
        dis2 = str(word[2])
        dis3 = str(word[3])
    elif repeat == 4 or repeat == 5 or repeat == 6:
        dis1 = str(word[3])
        dis2 = str(word[2])
        dis3 = str(word[1])
    if repeat == 1 or repeat == 4:
        new_amounts.pop(dis_row)
    current_amounts = new_amounts
    return dis_row, dis1, dis2, dis3, current_amounts

def weighted_rand(wordlist, current_amounts):
    weightings = []
    for amount in current_amounts:
        weightings.append(ceil(1000*((2/3)**current_amounts[amount])))
    word_as_a_list = random.choices(wordlist,weights=weightings, k=1)
    word = word_as_a_list[0]
    return word, current_amounts

finished = False
#dis1, dis2 and dis3 for the staggered
current_amounts = {}
if repeat == 1 or repeat == 2 or repeat == 3:
    for existing in initial_word_amounts_je:
        current_amounts[existing] = initial_word_amounts_je[existing]
elif repeat == 4 or repeat == 5 or repeat == 6:
    for existing in initial_word_amounts_je:
        current_amounts[existing] = initial_word_amounts_ej[existing]

while True:

    #if staggeryn == True:

    dis_row, dis1, dis2, dis3, current_amounts = stagger_question(repeat, words, current_amounts)
    print(dis1)
    input()
    print(dis2)
    input()
    print(dis3)
    inp = input("---------------------------how'd you go?")

    #adding 1 to the amounts if y is input hee hee
    
    if inp == 'y' or inp == 'z':
        if repeat <= 3:
            prev_amounts = ws[str('D'+str(dis_row))].value
            ws[str('D'+str(dis_row))] = ws[str('D'+str(dis_row))].value + 1
        elif repeat >= 4:
            prev_amounts = ws[str('E'+str(dis_row))].value
            ws[str('E'+str(dis_row))] = ws[str('E'+str(dis_row))].value + 1
        
        wb.save('VocabBook.xlsx')
        if repeat == 3 or repeat == 6:
            for wrd in words:
                if wrd[0] == dis_row:
                    words.remove(wrd)
                    break
            current_amounts.pop(dis_row)
    elif inp == 'n' or inp == 'x':
        if repeat <= 3:
            prev_amounts = ws[str('D'+str(dis_row))].value
            ws[str('D'+str(dis_row))] = ws[str('D'+str(dis_row))].value - 1
            if ws[str('D'+str(dis_row))].value > 0:
                ws[str('D'+str(dis_row))] = 0
        elif repeat >= 4:
            prev_amounts = ws[str('E'+str(dis_row))].value
            ws[str('E'+str(dis_row))] = ws[str('E'+str(dis_row))].value - 1
            if ws[str('E'+str(dis_row))].value > 0:
                ws[str('E'+str(dis_row))] = 0
        wb.save('VocabBook.xlsx')
            
    
    #re getting the swaggy amounts from the big lad list lad thingo lad yes
    row_num = 0
    for row in ws.values:
        row_num +=1
        if row[1] != None:
            initial_word_amounts_je[row_num] = row[3]
    for existing in current_amounts:
        current_amounts[existing] = initial_word_amounts_je[existing]
    
    current_amounts[1] = 'peepee poopoo'
    current_amounts.pop(1)

    if finished == True:
        break

    if len(current_amounts) < 2:
        finished = True

print()
print ("good study! have a good day :)")
input()
    
